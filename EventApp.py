import MyApp
import os
import wx
import openpyxl
import configparser

from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from time import sleep

import EventApp

infomart_file = "informart_list.xlsx"
config_file = os.getcwd() + "/config.ini"


class CEventApp(MyApp.MainFrame):

    def __init__(self, parent):
        MyApp.MainFrame.__init__(self, parent)

        self.grid_disp.AppendCols(2)
        self.grid_disp.SetColLabelValue(0, "商品番号")
        self.grid_disp.SetColLabelValue(1, "単価")

        is_file = os.path.isfile(infomart_file)

        if is_file:
            print(f"{infomart_file} is a file.")
        else:
            wb = openpyxl.Workbook()

            ws = wb.active

            # 書き込み処理
            ws["A1"] = "商品番号"
            ws["B1"] = "単価"
            wb.save(infomart_file)

        is_file = os.path.isfile(config_file)
        if is_file:
            id = Config.ReadConfig("USER", "id")
            pw = Config.ReadConfig("USER", "pw")
            self.text_ctrl_id.SetValue(id)
            self.text_ctrl_pw.SetValue(pw)
        else:
            Config.InitialConfig()  # 設定ファイル作成

    def OnRead(self, event):  # wxGlade: MainFrame.<event_handler>
        print("Event handler 'OnRead' not implemented!")

        wb = openpyxl.load_workbook(infomart_file)
        ws = wb.active

        row = 0
        for cells in tuple(ws.rows):
            col = 0
            for cell in cells:
                if cell.value == "":
                    break
                if row == 0:
                    continue
                elif row > self.grid_disp.NumberRows:
                    self.grid_disp.AppendRows(1)

                # print(cell.value)
                self.grid_disp.SetCellValue(
                    int(row)-1, int(col), str(cell.value))
                col = col + 1

            row = row + 1


        Csea = EventApp.CSubEventApp
        frame = Csea(None)
        frame.Show(True)


    def OnExec(self, event):  # wxGlade: MainFrame.<event_handler>
        print("Event handler 'OnExec' not implemented!")

        Config.WriteConfig("USER","id",self.text_ctrl_id.GetValue())
        Config.WriteConfig("USER","pw",self.text_ctrl_pw.GetValue())

        id = self.text_ctrl_id.GetValue()
        pw = self.text_ctrl_pw.GetValue()
        web = "https://www.infomart.co.jp/scripts/logon.asp"

        options = Options()
        # 画面を閉じない
        options.add_experimental_option('detach', True)
        # Chromeを起動
        driver = webdriver.Chrome(
            ChromeDriverManager().install(), options=options)

        # ログインページを開く
        driver.get(web)

        # ログオン処理
        # ID/PW入力
        driver.find_element(By.NAME, 'UID').send_keys(id)
        driver.find_element(By.NAME, 'PWD').send_keys(pw)

        driver.find_element(By.NAME, 'Logon').click()

        sleep(5)
        # トップページ
        driver.find_element(By.CLASS_NAME, 'mym-u-tanka').click()

        wx.MessageBox('得意先選択後、okボタンを押して続行してください。')

        # 転記
        row = self.grid_disp.NumberRows
        for i in range(0, row):
            if "" == self.grid_disp.GetCellValue(i, 0):
                break

            id_xpath = '//*[@id="item_private_code' + str(i) + '"]'
            pw_xpath = '//*[@id="prod_lot_price' + str(i) + '"]'

            driver.find_element(By.XPATH, id_xpath).send_keys(
                self.grid_disp.GetCellValue(i, 0))
            driver.find_element(By.XPATH, pw_xpath).send_keys(
                self.grid_disp.GetCellValue(i, 1))


class CSubEventApp(MyApp.SubFrame):

    def __init__(self, parent):
        MyApp.SubFrame.__init__(self, parent)

    def ClickList(self, event):  # wxGlade: SubFrame.<event_handler>
        print("Event handler 'ClickList' not implemented!")
        event.Skip()

# 設定ファイルクラス
class Config():
    
    def InitialConfig():

        config = configparser.RawConfigParser()

        section1 = 'USER'
        config.add_argument('--headless')
        config.add_section(section1)
        config.set(section1, 'id', '')
        config.set(section1, 'pw', '')

        with open(config_file, 'w') as file:
            config.write(file)

    def WriteConfig(section, palam, value):

        config = configparser.RawConfigParser()
        config.read(config_file)

        config.set(section, palam, value)

        with open(config_file, 'w') as file:
            config.write(file)

    def ReadConfig(section, palam):

        config = configparser.ConfigParser()
        config.read(config_file)

        return config.get(section, palam)
